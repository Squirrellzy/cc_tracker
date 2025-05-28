import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
import base64
import requests
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

st.set_page_config(page_title="CC Tracker – Indy", layout="wide")

# Style
st.markdown("""
    <style>
    .title-wrapper { text-align: center; margin-bottom: 2rem; }
    </style>
""", unsafe_allow_html=True)

# Dropdown options
options = ["", "Tracked", "Needs Tracked", "Pulley Noise", "Inspected"]
cc_list = [f"CC{i}" for i in range(1, 78)]

# Initialize session state
if "df" not in st.session_state:
    st.session_state.df = pd.DataFrame({
        "CC#": cc_list,
        "(A)-1": ["" for _ in cc_list],
        "2": ["" for _ in cc_list],
        "3": ["" for _ in cc_list],
        "4-(B)": ["" for _ in cc_list],
        "COMMENTS": ["" for _ in cc_list],
    })

st.markdown("<h2 class='title-wrapper'>Collection Conveyor Tracker – Indy</h2>", unsafe_allow_html=True)

# Add labels for dropdown columns
st.markdown("**(A)-1**   **2**   **3**   **4-(B)**")

# Display editable table with dropdowns and comment field only
edited_df = st.data_editor(
    # Disable sorting and make dropdowns open in one tap
    st.session_state.df,
    column_config={
        "(A)-1": st.column_config.SelectboxColumn("(A)-1", options=options),
        "2": st.column_config.SelectboxColumn("2", options=options),
        "3": st.column_config.SelectboxColumn("3", options=options),
        "4-(B)": st.column_config.SelectboxColumn("4-(B)", options=options),
        "COMMENTS": st.column_config.TextColumn("COMMENTS")
    },
    use_container_width=True,
    num_rows="fixed",
    hide_index=True,
    column_order=["CC#", "(A)-1", "2", "3", "4-(B)", "COMMENTS"],
    disabled=["CC#"],
    key="editor_no_sort",
    sort_by=None
)

# GitHub secrets
GITHUB_TOKEN = st.secrets["GITHUB_TOKEN"]
REPO_OWNER = st.secrets["REPO_OWNER"]
REPO_NAME = st.secrets["REPO_NAME"]
GITHUB_FILE = "CC Inspection Indy.xlsx"

def auto_format_worksheet(ws):
    tab = Table(displayName="InspectionLog", ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

def get_github_file():
    url = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/{GITHUB_FILE}"
    headers = {"Authorization": f"Bearer {GITHUB_TOKEN}", "Accept": "application/vnd.github+json"}
    resp = requests.get(url, headers=headers)
    if resp.status_code == 200:
        content = base64.b64decode(resp.json()["content"])
        sha = resp.json()["sha"]
        return BytesIO(content), sha
    return None, None

def push_to_github(buf, sha=None):
    b64 = base64.b64encode(buf.getvalue()).decode()
    url = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/{GITHUB_FILE}"
    headers = {"Authorization": f"Bearer {GITHUB_TOKEN}", "Accept": "application/vnd.github+json"}
    payload = {
        "message": f"Update sheet for {datetime.now().date()}",
        "content": b64,
        "branch": "main"
    }
    if sha: payload["sha"] = sha
    return requests.put(url, headers=headers, json=payload)

def save_and_upload(df):
    today = datetime.now().strftime("%Y-%m-%d")
    buf = BytesIO()
    file, sha = get_github_file()
    if file:
        book = load_workbook(file)
    else:
        from openpyxl import Workbook
        book = Workbook(); book.remove(book.active)
    if today in book.sheetnames:
        del book[today]
    ws = book.create_sheet(title=today)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    auto_format_worksheet(ws)
    book.save(buf)
    buf.seek(0)
    return push_to_github(buf, sha), buf

if st.button("Save to GitHub"):
    (resp, out_buf) = save_and_upload(edited_df)
    if resp.status_code in [200, 201]:
        st.success("✅ Workbook updated on GitHub!")
        st.session_state["download_buffer"] = out_buf
    else:
        st.error(f"❌ Failed to upload: {resp.json()}")
        st.session_state["download_buffer"] = None

if "download_buffer" in st.session_state and st.session_state["download_buffer"]:
    st.download_button("📥 Download This Version", st.session_state["download_buffer"], file_name="CC Inspection Indy.xlsx")
