
import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

st.set_page_config(page_title="CC Tracker – Indy (LAN)", layout="wide")

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    st.title("🔐 Login Required")
    username = st.text_input("Username")
    password = st.text_input("Password", type="password")
    if st.button("Login"):
        if username == "maint" and password == "mars":
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("Incorrect username or password.")
    st.stop()

options = ["", "Tracked", "Needs Tracked", "Pulley Noise", "Inspected"]
cc_list = [f"CC{i}" for i in range(1, 78)]

if "form_data" not in st.session_state:
    st.session_state.form_data = {}

st.title("Collection Conveyor Tracker – Indy (Offline Mode)")

for cc in cc_list:
    with st.container():
        st.subheader(cc)
        col1, col2, col3, col4, col5 = st.columns([1, 1, 1, 1, 3])
        with col1:
            a1 = st.selectbox("(A)-1", options, key=f"{cc}-a1")
        with col2:
            b2 = st.selectbox("2", options, key=f"{cc}-2")
        with col3:
            b3 = st.selectbox("3", options, key=f"{cc}-3")
        with col4:
            b4 = st.selectbox("4-(B)", options, key=f"{cc}-4")
        with col5:
            comment = st.text_input("COMMENTS", key=f"{cc}-comment")
        st.session_state.form_data[cc] = [a1, b2, b3, b4, comment]

def auto_format_worksheet(ws):
    tab = Table(displayName="InspectionLog", ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

def save_to_excel():
    today = datetime.now().strftime("%Y-%m-%d")
    filename = "CC Inspection Indy.xlsx"
    if os.path.exists(filename):
        book = load_workbook(filename)
    else:
        book = Workbook()
        book.remove(book.active)
    if today in book.sheetnames:
        del book[today]
    ws = book.create_sheet(title=today)
    ws.append(["CC#", "(A)-1", "2", "3", "4-(B)", "COMMENTS"])
    for cc in cc_list:
        ws.append([cc] + st.session_state.form_data[cc])
    auto_format_worksheet(ws)
    buf = BytesIO()
    book.save(buf)
    book.save(filename)
    buf.seek(0)
    return buf

if st.button("💾 Save to Local File"):
    buffer = save_to_excel()
    st.success("Saved to 'CC Inspection Indy.xlsx'")
    st.session_state.download_buffer = buffer

if "download_buffer" not in st.session_state:
    st.session_state.download_buffer = save_to_excel()

st.download_button("📥 Download Workbook", st.session_state.download_buffer, file_name="CC Inspection Indy.xlsx")
